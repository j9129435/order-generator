import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import date

# --- 頁面設定 ---
st.set_page_config(page_title="報價單生成系統 (含成本分析)", layout="wide", page_icon="📊")

st.title("📊 報價單與成本分析生成器")
st.info("💡 支援填寫成本與供應商資訊，系統將自動計算利潤並填入 Excel 隱藏欄位。")

# --- 1. 側邊欄：業務與系統設定 (綠色區塊) ---
st.sidebar.header("1. 業務資訊 (綠色區塊)")
sales_name = st.sidebar.text_input("承辦業務", "陳書豪 (台中業務部)")
sales_mobile = st.sidebar.text_input("業務手機", "0934-290929")
sales_line = st.sidebar.text_input("LINE ID", "powerhao")
sales_email = st.sidebar.text_input("電子信箱", "powerhao.chen@fongcon.com.tw")

st.sidebar.divider()
st.sidebar.header("2. 範本設定")
uploaded_template = st.sidebar.file_uploader("上傳 Excel 範本", type=["xlsx"])
# 若無上傳，預設讀取同目錄下的 template.xlsx
template_source = uploaded_template if uploaded_template else "template.xlsx"

# --- 2. 主畫面：客戶資訊 (紅色區塊) ---
st.header("📝 客戶基本資料 (紅色區塊)")

col1, col2 = st.columns(2)

with col1:
    customer_name = st.text_input("客戶名稱", "康葳國際生醫有限公司")
    department = st.text_input("隸屬部門", "")
    contact_person = st.text_input("聯 絡 人", "邱惠微 Vivi Chiu")
    phone = st.text_input("公司電話", "04-22360750")
    fax = st.text_input("公司傳真", "04-22360720")

with col2:
    mobile = st.text_input("行動電話", "0927-701927")
    tax_id = st.text_input("統一編號", "45883386")
    address = st.text_input("公司地址", "台中市北屯區崇德路二段130號6樓")
    email = st.text_input("E - mail", "twou1635@gmail.com")
    quotation_date = st.date_input("報價日期", date.today())

# --- 3. 商品明細 (藍色+黃色區塊) ---
st.header("📦 商品與成本明細 (藍色/黃色區塊)")
st.caption("請在表格中輸入商品售價 (藍色) 與 內部成本 (黃色)，系統會自動計算利潤。")

# 預設資料表格
if "df_items" not in st.session_state:
    st.session_state.df_items = pd.DataFrame(
        [
            {
                "廠牌": "HP", "型號": "PRO400G9M", "規格": "處理器:i5-14500 / 32G / 1TB SSD", 
                "數量": 1, "售價(單價)": 31000, "成本(單價)": 22500, "供應商": "聯強"
            },
            {
                "廠牌": "", "型號": "", "規格": "記憶體: 32G DDR5", 
                "數量": 1, "售價(單價)": 0, "成本(單價)": 2600, "供應商": "庫存"
            },
        ]
    )

# 顯示可編輯表格
edited_df = st.data_editor(
    st.session_state.df_items,
    num_rows="dynamic",
    column_config={
        "數量": st.column_config.NumberColumn(format="%d"),
        "售價(單價)": st.column_config.NumberColumn(format="$%d", label="🔵 售價 (單價)"),
        "成本(單價)": st.column_config.NumberColumn(format="$%d", label="🟡 成本 (單價)"),
        "供應商": st.column_config.TextColumn(label="🟡 供應商"),
    },
    use_container_width=True
)

# --- 4. 核心邏輯：寫入 Excel ---
def generate_excel(template_src, data, items_df, sales_data):
    try:
        wb = openpyxl.load_workbook(template_src)
        ws = wb.active
        
        # ==========================================
        #⚠️ 座標設定 (根據 219康葳...xlsx)
        # ==========================================
        
        # --- 紅色區塊 (客戶) ---
        ws['B9'] = data['customer_name']
        ws['B10'] = data['department']
        ws['B11'] = data['contact_person']
        ws['B12'] = data['phone']
        ws['B13'] = data['fax']
        ws['B14'] = data['mobile']
        ws['B15'] = data['tax_id']
        ws['B16'] = data['address']
        ws['B17'] = data['email']
        
        # --- 綠色區塊 (業務) ---
        ws['B38'] = sales_data['name']
        ws['B39'] = sales_data['mobile']
        ws['B40'] = sales_data['line']
        ws['B41'] = sales_data['email']
        # 報價日期 (B42) 與 客戶簽名欄位旁
        ws['B42'] = data['quotation_date'] 

        # --- 藍色 & 黃色區塊 (商品) ---
        start_row = 20  # 商品起始列
        
        total_price = 0
        total_cost = 0
        
        for index, row in items_df.iterrows():
            r = start_row + index
            
            # 處理空值，避免計算錯誤
            qty = row['數量'] if pd.notnull(row['數量']) else 0
            price = row['售價(單價)'] if pd.notnull(row['售價(單價)']) else 0
            cost = row['成本(單價)'] if pd.notnull(row['成本(單價)']) else 0
            
            subtotal_price = qty * price
            subtotal_cost = qty * cost
            
            total_price += subtotal_price
            total_cost += subtotal_cost
            
            # 寫入儲存格
            ws[f'A{r}'] = row['廠牌']
            ws[f'B{r}'] = row['型號']
            ws[f'C{r}'] = row['規格']
            ws[f'D{r}'] = qty
            ws[f'E{r}'] = price
            ws[f'F{r}'] = subtotal_price # 售價小計 (藍)
            
            ws[f'G{r}'] = cost           # 成本單價 (黃)
            ws[f'H{r}'] = subtotal_cost  # 成本小計 (黃)
            ws[f'I{r}'] = row['供應商']   # 供應商 (黃)

        # --- 橘色 & 統計區塊 ---
        # 售價統計 (顯示給客戶)
        tax_rate = 0.05
        tax_amount = total_price * tax_rate
        grand_total = total_price + tax_amount
        
        ws['F29'] = total_price    # 合計(未稅)
        ws['F30'] = tax_amount     # 營業稅
        ws['F31'] = grand_total    # 總計金額 (橘色)

        # 成本與利潤統計 (內部查看 - H欄)
        total_profit = total_price - total_cost
        profit_margin = (total_profit / total_price) if total_price > 0 else 0
        
        ws['H29'] = total_cost     # 總成本
        ws['H30'] = total_profit   # 總毛利
        ws['H31'] = profit_margin  # 毛利率

        # 輸出檔案到記憶體
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        st.error(f"Excel 處理發生錯誤: {e}")
        return None

# --- 5. 生成按鈕 ---
st.divider()
col_btn, col_info = st.columns([1, 3])

with col_btn:
    generate_btn = st.button("🚀 生成報價單", type="primary")

# 這裡的邏輯控制非常重要，縮排必須正確
if generate_btn:
    # 1. 整理資料
    customer_data = {
        "customer_name": customer_name,
        "department": department,
        "contact_person": contact_person,
        "phone": phone,
        "fax": fax,
        "mobile": mobile,
        "tax_id": tax_id,
        "address": address,
        "email": email,
        "quotation_date": quotation_date
    }
    
    sales_data = {
        "name": sales_name,
        "mobile": sales_mobile,
        "line": sales_line,
        "email": sales_email
    }
    
    # 2. 執行生成
    excel_file = generate_excel(template_source, customer_data, edited_df, sales_data)
    
    # 3. 如果成功生成，顯示下載按鈕
    if excel_file:
        file_name = f"報價單_{customer_name}_{date.today()}.xlsx"
        st.success(f"成功生成！請下載檔案。")
        st.download_button(
            label="📥 下載 Excel 檔案",
            data=excel_file,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )